"""Optional XLSX projection for structured report tables.

The exporter deliberately follows :mod:`deep_research.report.csv`: one
``TableBlock`` is selected, values are emitted in their human-facing form,
and citations, protocol notes, missing cells, and references stay visible in
the workbook.  ``openpyxl`` is imported only when an export is requested so
it remains an optional dependency and cannot make API startup fail.
"""

from __future__ import annotations

import io
import re
from typing import Any

from .csv import (
    CsvTableNotFoundError,
    CsvTableSelectionError,
    _cell_value,
    _column_header,
    _select_table,
)
from .document import ReportDocument, TableBlock


class XlsxExportError(ValueError):
    """Base error for an invalid structured-table XLSX request."""


class XlsxDependencyError(XlsxExportError):
    """Raised when the optional ``openpyxl`` package is not installed."""


class XlsxTableNotFoundError(XlsxExportError):
    """Raised when a requested table id is not present in the document."""


class XlsxTableSelectionError(XlsxExportError):
    """Raised when a document contains multiple tables without a selection."""


# Excel worksheet names may not contain these characters and are capped at 31
# characters.  Keep the title deterministic so downloads are easy to inspect.
_INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")
_MAX_CELL_TEXT = 32_767


def render_xlsx(
    document: ReportDocument,
    *,
    table_id: str | None = None,
    include_references: bool = True,
) -> bytes:
    """Render one structured report table as an XLSX byte string.

    A report with no table still yields a valid empty workbook.  This mirrors
    the streaming-friendly CSV endpoint (which yields an empty body) while
    ensuring that an ``.xlsx`` download can always be opened by spreadsheet
    software.  Cell contents are stored as text, including values beginning
    with ``=``/``+``/``-``/``@``; report data must never be interpreted as an
    Excel formula.
    """

    Workbook, styles = _load_openpyxl()
    table = _select_xlsx_table(document, table_id)

    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = _sheet_title(table)
    worksheet.freeze_panes = "B2"

    # Keep all values text so citations and footnote markers cannot be lost or
    # accidentally evaluated as formulas by Excel.
    header_fill = styles.PatternFill("solid", fgColor="1F4E78")
    header_font = styles.Font(color="FFFFFF", bold=True)
    note_fill = styles.PatternFill("solid", fgColor="FFF2CC")
    border = styles.Border(
        bottom=styles.Side(style="thin", color="D9E2F3"),
    )

    if table is not None and table.columns:
        headers = ["对象", *(_column_header(column) for column in table.columns)]
        for column_index, value in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=column_index)
            _set_text(cell, value)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = styles.Alignment(horizontal="center", vertical="top", wrap_text=True)

        for row_index, row in enumerate(table.rows, 2):
            label = row.label
            if row.citation:
                label = f"{label} [{row.citation}]"
            values = [label, *(_cell_value(row.cell(column.key)) for column in table.columns)]
            for column_index, value in enumerate(values, 1):
                cell = worksheet.cell(row=row_index, column=column_index)
                _set_text(cell, value)
                cell.border = border
                cell.alignment = styles.Alignment(vertical="top", wrap_text=True)

        worksheet.auto_filter.ref = worksheet.dimensions
        _set_column_widths(worksheet, headers, table)
        next_row = len(table.rows) + 3
        width = len(headers)

        if table.notes:
            _write_section_heading(worksheet, next_row, "口径脚注", width, note_fill, styles)
            next_row += 1
            for index, note in enumerate(table.notes, 1):
                _write_full_width_row(worksheet, next_row, f"[{index}] {note}", width, styles)
                next_row += 1
            next_row += 1

        if table.caption:
            _write_section_heading(worksheet, next_row, "说明", width, note_fill, styles)
            next_row += 1
            _write_full_width_row(worksheet, next_row, table.caption, width, styles)
            next_row += 2

        if include_references and document.references:
            _write_section_heading(worksheet, next_row, "参考来源", width, note_fill, styles)
            next_row += 1
            for reference in document.references:
                _write_full_width_row(
                    worksheet,
                    next_row,
                    f"[{reference.index}] {reference.render()}",
                    width,
                    styles,
                )
                next_row += 1
    else:
        _set_text(worksheet["A1"], "暂无表格")
        worksheet["A1"].font = styles.Font(bold=True)
        worksheet.column_dimensions["A"].width = 18

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _load_openpyxl() -> tuple[Any, Any]:
    """Load openpyxl lazily and expose a stable error for minimal installs."""

    try:
        from openpyxl import Workbook, styles
    except ImportError as exc:  # pragma: no cover - exercised with an extra-less env
        raise XlsxDependencyError(
            "XLSX export requires the optional dependency 'openpyxl'; install the 'xlsx' extra"
        ) from exc
    return Workbook, styles


def _select_xlsx_table(document: ReportDocument, table_id: str | None) -> TableBlock | None:
    """Map the CSV selector's errors to the XLSX-specific public types."""

    tables = [block for block in document.blocks if isinstance(block, TableBlock)]
    try:
        return _select_table(tables, table_id)
    except CsvTableNotFoundError as exc:
        raise XlsxTableNotFoundError(str(exc)) from exc
    except CsvTableSelectionError as exc:
        raise XlsxTableSelectionError(str(exc)) from exc


def _sheet_title(table: TableBlock | None) -> str:
    if table is None:
        return "Report"
    raw = table.title.strip() or table.id.strip() or "Report"
    title = _INVALID_SHEET_CHARS.sub("_", raw).strip("'")[:31]
    return title or "Report"


def _set_text(cell: Any, value: str) -> None:
    """Store a string explicitly, preventing formula interpretation."""

    # Excel's OOXML text-cell limit is 32,767 UTF-16 code units.  Truncating
    # only at this final serialization boundary keeps the in-memory document
    # and CSV export lossless while avoiding a corrupt workbook for pathological
    # model output.
    text = str(value)
    if len(text) > _MAX_CELL_TEXT:
        text = text[:_MAX_CELL_TEXT]
    cell.value = text
    cell.data_type = "s"


def _write_section_heading(
    worksheet: Any,
    row: int,
    value: str,
    width: int,
    fill: Any,
    styles: Any,
) -> None:
    cell = worksheet.cell(row=row, column=1)
    _set_text(cell, value)
    cell.fill = fill
    cell.font = styles.Font(bold=True)
    if width > 1:
        worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)


def _write_full_width_row(worksheet: Any, row: int, value: str, width: int, styles: Any) -> None:
    cell = worksheet.cell(row=row, column=1)
    _set_text(cell, value)
    cell.alignment = styles.Alignment(vertical="top", wrap_text=True)
    if width > 1:
        worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)


def _set_column_widths(worksheet: Any, headers: list[str], table: TableBlock) -> None:
    """Choose bounded widths that keep long evidence text readable."""

    from openpyxl.utils import get_column_letter

    for index, header in enumerate(headers, 1):
        values = [header]
        for row in table.rows:
            if index == 1:
                value = row.label
            else:
                column = table.columns[index - 2]
                value = _cell_value(row.cell(column.key))
            values.append(value)
        # Evidence cells wrap instead of expanding a sheet to an unusable
        # width.  Include their short values in sizing so ordinary tables do
        # not look needlessly compressed.
        width = min(max(len(str(value)) for value in values), 42)
        worksheet.column_dimensions[get_column_letter(index)].width = max(12, width + 2)


__all__ = [
    "XlsxDependencyError",
    "XlsxExportError",
    "XlsxTableNotFoundError",
    "XlsxTableSelectionError",
    "render_xlsx",
]
